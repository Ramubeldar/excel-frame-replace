import cv2
import moviepy.editor as mp
import os
from openpyxl import load_workbook


# defining the function
def read_excel_row(row_num, excel_file):
    wb = load_workbook(filename=excel_file)
    worksheet = wb.active

    Input_Video = []
    for col in range(1, 2):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Input_Video.append(cell_value)

    Input_Image = []
    for col in range(2, 3):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Input_Image.append(cell_value)

    Start_Time = []
    for col in range(3, 4):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Start_Time.append(cell_value)

    Duration = []
    for col in range(4, 5):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Duration.append(cell_value)

    Output_Folder = []
    for col in range(5, 6):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Output_Folder.append(cell_value)

    input_video_path = Input_Video[0]

    # Path to the replacement picture
    replacement_image_path = Input_Image[0]

    # Output video path
    output_video_path = 'output_video.mp4'

    # Time in seconds to start replacing the frame
    start_time = Start_Time[0]

    # Duration in seconds for which to replace the frame
    duration = Duration[0]

    # Load the input video
    video = cv2.VideoCapture(input_video_path)

    # Read the first frame to get frame properties
    ret, frame = video.read()
    frame_height, frame_width, _ = frame.shape

    # Calculate frame numbers to start and stop replacing frames
    frame_rate = video.get(cv2.CAP_PROP_FPS)
    start_frame = int(start_time * frame_rate)
    end_frame = int((start_time + duration) * frame_rate)

    # Load the replacement image
    replacement_image = cv2.imread(replacement_image_path)

    # Create a new video writer object for the output video
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    output_video = cv2.VideoWriter(output_video_path, fourcc, frame_rate, (frame_width, frame_height))

    # Iterate through the frames and replace the desired frames
    frame_count = 0
    while True:
        ret, frame = video.read()
        if not ret:
            break

        if frame_count >= start_frame and frame_count <= end_frame:
            # Resize the replacement image to match the frame size
            resized_image = cv2.resize(replacement_image, (frame_width, frame_height))

            # Replace the frame with the resized replacement image
            frame = resized_image

        # Save the modified frame to the output video
        output_video.write(frame)

        frame_count += 1

    # Release resources
    video.release()
    output_video.release()

    # Path to the source video
    source_video_path = Input_Video[0]

    # Path to save the extracted audio
    audio_output_path = 'extracted_audio.mp3'

    # Path to the target video
    target_video_path = 'output_video.mp4'

    # Path to save the final video with added audio
    final_output_path = 'final_video.mp4'

    # Extract audio from the source video
    video_clip = mp.VideoFileClip(source_video_path)
    audio_clip = video_clip.audio
    audio_clip.write_audiofile(audio_output_path)

    # Load the target video
    target_video_clip = mp.VideoFileClip(target_video_path)

    # Load the extracted audio
    extracted_audio_clip = mp.AudioFileClip(audio_output_path)

    # Set the extracted audio to the target video
    final_video = target_video_clip.set_audio(extracted_audio_clip)

    # create the subfolder within the top-level folder
    # Create a VideoWriter object
    output_folder = Output_Folder[0]
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, "output.mp4")
    # Write the final video with added audio
    final_video.write_videofile(output_path, codec='libx264')

    # delete the extracted_audio.mp3 file from disk
    file_path = "extracted_audio.mp3"
    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        pass

    print("Modified video saved successfully!")


# reading the exel file
excel_file_sheets = "data.xlsx"
workbook = load_workbook("data.xlsx")

# Select the active worksheet
worksheet = workbook.active

# Find the number of rows with data in excel sheet
num_rows = 0
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    if any(cell.value for cell in row):
        num_rows += 1

for row_number in range(2, num_rows + 1):
    read_excel_row(row_number, excel_file_sheets)
